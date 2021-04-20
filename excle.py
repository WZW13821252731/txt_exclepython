import openpyxl
import codecs
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import re
 
def txt_to_xlsx(filename,outfile):
 
    fr = codecs.open(filename,encoding='UTF-8')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws1 = wb.create_sheet()
    ws1.title = 'Sheet1'
    row = 0  #行
    for line in fr:
        row +=1
        line = line.strip() #strip() 方法用于移除字符串头尾指定的字符（默认为空格或换行符）或字符序列。
                             #注意：该方法只能删除开头或是结尾的字符，不能删除中间部分的字符。
        line = line.split('\t')#通过指定分隔符对字符串进行切片，如果参数 num 有指定值，则分隔 num+1 个子字符串
        #line = line.split(":")
        col = 0
        # h = len(line)
        # print(h)
        for j in range(len(line)):
            col +=1
            #print (line[j])
            ws.cell(column = col,row = row,value = line[j].format(get_column_letter(col)))
    ws1.cell(column = 1,row =1,value= "温度" )
    ws1.cell(column = 9,row =1,value= "湿度" )
    for i  in range(8):
        ws1.cell(column = i+1,row =2,value= i+1)
    ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    for i in range(9,17):
        ws1.cell(column = i,row =2,value= i-8)
    ws1.merge_cells(start_row=1, start_column=9, end_row=1, end_column=16)
    print(type(ws.cell(row=1,column=2).value))
    print(re.findall(r"\d+\.?\d*",ws.cell(row=1,column=2).value)[0])
    # ws1.cell(column = 2,row =3,value=re.findall(r"\d+\.?\d*",ws.cell(row=1,column=2).value)[0])

    for j in range(int(ws.max_row/8)):
        for i in range(8):
            if i !=7:
                if i > 2:
                    ws1.cell(column = i+2,row =3+j,value=float(re.findall(r"\d+\.?\d*",ws.cell(row=i+1+j*8,column=2).value)[0]))
                    ws1.cell(column = i+10,row =3+j,value=float(re.findall(r"\d+\.?\d*",ws.cell(row=i+1+j*8,column=3).value)[0]))
                else:
                    ws1.cell(column = i+1,row =3+j,value=float(re.findall(r"\d+\.?\d*",ws.cell(row=i+1+j*8,column=2).value)[0]))
                    ws1.cell(column = i+9,row =3+j,value=float(re.findall(r"\d+\.?\d*",ws.cell(row=i+1+j*8,column=3).value)[0]))
                    
    
    wb.save(outfile)
#读取xlsx内容    
def read_xlsx(filename):
    #载入文件
    wb = openpyxl.load_workbook(filename)
    #获取Sheet1工作表
    ws = wb.get_sheet_by_name('Sheet1')
    #按行读取
    for row in ws.rows:
        for cell in row:
            print (cell.value)
    #按列读
    for col in ws.columns:
        for cell in col:
            print (cell.value)
 
if __name__=='__main__':
    inputfileTxt = 'te.txt'
    outfileExcel = 'temperature.xlsx'
    txt_to_xlsx(inputfileTxt,outfileExcel)
    #read_xlsx(outfileExcel)



'''
data = []
f = open("temperatureHumidity.txt",encoding='UTF-8')   #设置文件对象
for line in f:            #直到读取完文件
    #print(line)
    data.append(line)
print(data)

f.close() #关闭文件
'''








































'''
import openpyxl
import codecs
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
 
def txt_to_xlsx(filename,outfile):
 
    fr = codecs.open(filename,encoding='UTF-8')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws1 = wb.create_sheet()
    ws1.title = 'Sheet1'
    row = 0  #行
    for line in fr:
        row +=1
        line = line.strip() #strip() 方法用于移除字符串头尾指定的字符（默认为空格或换行符）或字符序列。
                             #注意：该方法只能删除开头或是结尾的字符，不能删除中间部分的字符。
        line = line.split('\t')#通过指定分隔符对字符串进行切片，如果参数 num 有指定值，则分隔 num+1 个子字符串
        col = 0
        # h = len(line)
        # print(h)
        for j in range(len(line)):
            col +=1
            #print (line[j])
            ws.cell(column = col,row = row,value = line[j].format(get_column_letter(col)))
    ws1.cell(column = 1,row =1,value= "温度" )
    ws1.cell(column = 9,row =1,value= "湿度" )
    for i  in range(8):
        ws1.cell(column = i+1,row =2,value= i+1)
    ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    for i in range(9,17):
        ws1.cell(column = i,row =2,value= i-8)
    ws1.merge_cells(start_row=1, start_column=9, end_row=1, end_column=16)
    print(ws.cell(row=1,column=2).value)
    for i in range(8):
        if i > 2:
             ws1.cell(column = i+2,row =3,value=ws.cell(row=i+1,column=2).value)
        else:
             ws1.cell(column = i+1,row =3,value=ws.cell(row=i+1,column=2).value)
    wb.save(outfile)
#读取xlsx内容    
def read_xlsx(filename):
    #载入文件
    wb = openpyxl.load_workbook(filename)
    #获取Sheet1工作表
    ws = wb.get_sheet_by_name('Sheet1')
    #按行读取
    for row in ws.rows:
        for cell in row:
            print (cell.value)
    #按列读
    for col in ws.columns:
        for cell in col:
            print (cell.value)
 
if __name__=='__main__':
    inputfileTxt = 'te.txt'
    outfileExcel = 'temperature.xlsx'
    txt_to_xlsx(inputfileTxt,outfileExcel)
    #read_xlsx(outfileExcel)




data = []
f = open("temperatureHumidity.txt",encoding='UTF-8')   #设置文件对象
for line in f:            #直到读取完文件
    #print(line)
    data.append(line)
print(data)

f.close() #关闭文件
'''
